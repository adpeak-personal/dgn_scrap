import random
import threading
import time
import subprocess
from datetime import datetime  # , timedelta
import sys
import os
import requests
import json
import re
import pyautogui as pg
import ctypes
from playwright.sync_api import sync_playwright
from openpyxl import load_workbook
import msoffcrypto
import io
from tkinter import *
import shutil

# import pyperclip
# import pygetwindow as gw
# from pywinauto import Desktop
# import clipboard as cb
# from ppadb.client import Client as AdbClient
# import keyboard
# from tkinter import ttk
# import winsound as ws
# import winsound as sd
# from bs4 import BeautifulSoup as bs
# from pathlib import Path
# from typing import Optional
# from pyparsing import And
# import getpass
# from PIL import Image
# import os  # 위에서 이미 import함
# from openai import OpenAI


CHROME_USER_DATA_DIR = os.path.expandvars(r"%LOCALAPPDATA%\Google\Chrome\User Data")
TEMP_DIR = os.path.join(os.getenv('TEMP'), 'daangn_dl')
_BASE_DIR = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__))
LAST_CHECK_FILE = os.path.join(_BASE_DIR, 'last_check_time.txt')

def parse_mail_date(date_str):
    """오전/오후 HH:MM 또는 MM.DD HH:MM (또는 MM. DD HH:MM) 형식 → datetime"""
    date_str = date_str.strip()
    now = datetime.now()
    if '오전' in date_str or '오후' in date_str:
        is_pm = '오후' in date_str
        time_part = re.sub(r'오전|오후', '', date_str).strip()
        h, m = map(int, time_part.split(':'))
        if is_pm and h != 12:
            h += 12
        elif not is_pm and h == 12:
            h = 0
        return now.replace(hour=h, minute=m, second=0, microsecond=0)
    else:
        # "06.11 12:12", "06. 10 23:40", "06/12 13:15" 형식 모두 처리
        normalized = re.sub(r'\s*[./]\s*', '/', date_str)  # 구분자를 /로 통일
        parts = normalized.split()
        month, day = map(int, parts[0].split('/'))
        if len(parts) >= 2:
            h, m = map(int, parts[1].split(':'))
        else:
            h, m = 0, 0
        return now.replace(month=month, day=day, hour=h, minute=m, second=0, microsecond=0)

def read_last_check_time(key='dgnmail', default='오전 11:00'):
    seed = {'dgnmail': '오전 11:00'}
    if not os.path.exists(LAST_CHECK_FILE):
        with open(LAST_CHECK_FILE, 'w', encoding='utf-8') as f:
            json.dump(seed, f, ensure_ascii=False, indent=2)
        return seed.get(key, default)
    try:
        with open(LAST_CHECK_FILE, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return data.get(key, default)
    except json.JSONDecodeError:
        # 기존 plain text → JSON으로 마이그레이션
        with open(LAST_CHECK_FILE, 'r', encoding='utf-8') as f:
            old_value = f.read().strip()
        data = {'dgnmail': old_value}
        with open(LAST_CHECK_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        return data.get(key, default)

def write_last_check_time(time_str=None, key='dgnmail'):
    if time_str is None:
        now = datetime.now()
        time_str = f"{now.month:02d}.{now.day:02d} {now.hour:02d}:{now.minute:02d}"
    if os.path.exists(LAST_CHECK_FILE):
        try:
            with open(LAST_CHECK_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
        except json.JSONDecodeError:
            data = {}
    else:
        data = {}
    data[key] = time_str
    with open(LAST_CHECK_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"기준 시간 저장 [{key}]: {time_str}")

def clear_temp_dir():
    if os.path.exists(TEMP_DIR):
        shutil.rmtree(TEMP_DIR)
    os.makedirs(TEMP_DIR)
    print(f"임시폴더 초기화: {TEMP_DIR}")

def read_excel_with_password(path, password='1234qwer'):
    with open(path, 'rb') as f:
        office_file = msoffcrypto.OfficeFile(f)
        office_file.load_key(password=password)
        decrypted = io.BytesIO()
        office_file.decrypt(decrypted)
    wb = load_workbook(decrypted)
    ws = wb.active
    return list(ws.iter_rows(values_only=True))

def get_profiles() -> dict[str, str]:
    """크롬 프로필 목록 반환 {표시 이름: 폴더명}"""
    local_state_path = os.path.join(CHROME_USER_DATA_DIR, "Local State")
    try:
        with open(local_state_path, encoding="utf-8") as f:
            data = json.load(f)
        cache = data.get("profile", {}).get("info_cache", {})
        return {info.get("name", folder): folder for folder, info in cache.items()}
    except Exception:
        return {"Default": "Default"}


def delay(a=1.0, b=1.5):
    time.sleep(random.uniform(a, b))

def back_to_main(new_tab, main_page):
    new_tab.close()
    delay()
    main_page.bring_to_front()
    main_page.reload()
    main_page.wait_for_selector('.mail_item')
    print("두번째 탭 닫고 첫번째 탭으로 이동 + 새로고침 완료")

def close_if_extra(context, new_page):
    if len(context.pages) > 2:
        new_page.close()
        print("여분 탭 닫음")

    
